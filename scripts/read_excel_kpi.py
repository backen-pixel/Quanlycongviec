"""Read both Excel files and dump their content for analysis."""
import sys
import io
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

files = [
    r"C:\Users\Admin\Downloads\CRM_Setup_QuyTrinh_TuBep.xlsx",
    r"C:\Users\Admin\Downloads\KPI_CRM_SalesAdmin_Deal_TuBep.xlsx",
]

for fp in files:
    print("=" * 100)
    print(f"FILE: {fp}")
    print("=" * 100)
    wb = load_workbook(fp, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- SHEET: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                print(" | ".join(cells))
        print()
